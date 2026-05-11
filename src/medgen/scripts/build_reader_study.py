"""Build a blind reader-study set (real vs synthetic) for radiologist evaluation.

Randomly samples N real volumes and N synthetic volumes (bravo + seg only),
normalizes them consistently, gives each a random 1..2N case ID, and writes:

  <output-dir>/volumes/<id>/bravo.nii.gz   (normalized to [0, 1])
  <output-dir>/volumes/<id>/seg.nii.gz     (binary segmentation, unchanged)
  <output-dir>/answer_key.csv              (case_id, true_label, source)   ← DO NOT SHARE
  <output-dir>/radiologist_form.csv        (blank form for the reader)     ← SHARE THIS

By default both sides are brain-mask zeroed so background appearance can't
be used as a tell. Disable with --no-brain-mask if the source dirs are
already consistent.

Usage:
    python -m medgen.scripts.build_reader_study \\
        --real-root /path/to/brainmetshare-3 --real-split train \\
        --synth-dir /path/to/generated/seg_candidates_525/exp48c_handoff_exp32 \\
        --output-dir runs/reader_study/exp48c_handoff_exp32 \\
        --n-real 50 --n-synth 50 --seed 42
"""
from __future__ import annotations

import argparse
import csv
import logging
import shutil
import sys
from pathlib import Path

import nibabel as nib
import numpy as np

logger = logging.getLogger(__name__)


def _find_subjects(root: Path) -> list[Path]:
    """Return all subdirs under root that have both bravo.nii.gz and seg.nii.gz."""
    if not root.is_dir():
        raise FileNotFoundError(root)
    out = []
    for sub in sorted(root.iterdir()):
        if not sub.is_dir():
            continue
        if (sub / "bravo.nii.gz").is_file() and (sub / "seg.nii.gz").is_file():
            out.append(sub)
    return out


def _find_real_subjects_across_splits(
    root: Path, splits: list[str]
) -> list[Path]:
    """Collect subjects across multiple splits, dedup by subject directory name.

    Stanford BrainMetShare has 156 unique patients spread across splits with
    overlap (e.g. test_new ⊂ test1). Walking all splits and deduping by the
    leaf directory name (subject ID) gives the full 156-patient pool.
    """
    seen: dict[str, Path] = {}
    for split in splits:
        split_dir = root / split
        if not split_dir.is_dir():
            logger.warning("split dir missing: %s — skipping", split_dir)
            continue
        for subj in _find_subjects(split_dir):
            # First-come wins on duplicates; doesn't matter which path —
            # files for the same subject are bit-identical across splits.
            seen.setdefault(subj.name, subj)
    return sorted(seen.values(), key=lambda p: p.name)


def _normalize_bravo(vol: np.ndarray) -> np.ndarray:
    """Min-max normalize to [0, 1]."""
    vmin, vmax = float(vol.min()), float(vol.max())
    if vmax > vmin:
        return ((vol - vmin) / (vmax - vmin)).astype(np.float32)
    return vol.astype(np.float32)


def _apply_brain_mask(
    vol: np.ndarray, threshold: float = 0.05, dilate_pixels: int = 2
) -> np.ndarray:
    """Zero outside brain — same logic as generate.py's mask_outside_brain step.

    Uses the project's create_brain_mask so behaviour matches what was
    applied to synth volumes at generation time.
    """
    from medgen.metrics.brain_mask import create_brain_mask
    mask = create_brain_mask(
        vol, threshold=threshold, fill_holes=True, dilate_pixels=dilate_pixels
    )
    return (vol * mask.astype(vol.dtype)).astype(np.float32)


def _save_nifti_like(
    data: np.ndarray, reference_img: nib.Nifti1Image, dst: Path
) -> None:
    """Save `data` as a NIfTI at `dst`, preserving affine + header from ref."""
    dst.parent.mkdir(parents=True, exist_ok=True)
    out = nib.Nifti1Image(data, reference_img.affine, reference_img.header)
    nib.save(out, str(dst))


def _write_xlsx_form(path: Path, case_ids: list[str], n_real: int, n_synth: int) -> bool:
    """Write an Excel form with dropdowns, frozen header, alternating row colors.

    Returns True on success, False if openpyxl is unavailable (CSV fallback).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        logger.warning(
            "openpyxl not installed — skipping XLSX form. "
            "Install with: pip install openpyxl"
        )
        return False

    wb = Workbook()

    # ── Sheet 1: Reader Form (active on open) ────────────────────────────────
    ws = wb.active
    ws.title = "Reader Form"

    # Header row
    headers = ["case_id", "prediction", "confidence", "comments"]
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 10  # case_id
    ws.column_dimensions["B"].width = 14  # prediction
    ws.column_dimensions["C"].width = 14  # confidence
    ws.column_dimensions["D"].width = 70  # comments

    # Freeze the header so it stays visible while scrolling
    ws.freeze_panes = "A2"

    # Case rows with alternating fill for readability
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    centered = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, case_id in enumerate(case_ids, start=2):
        c = ws.cell(row=row_idx, column=1, value=case_id)
        c.alignment = centered
        ws.cell(row=row_idx, column=2).alignment = centered
        ws.cell(row=row_idx, column=3).alignment = centered
        ws.cell(row=row_idx, column=4).alignment = left_wrap
        if row_idx % 2 == 0:
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill
        ws.row_dimensions[row_idx].height = 24

    # ── Dropdowns ────────────────────────────────────────────────────────────
    last_row = len(case_ids) + 1

    dv_pred = DataValidation(
        type="list",
        formula1='"real,synthetic,unsure"',
        allow_blank=True,
        showDropDown=False,  # show the dropdown arrow
        errorTitle="Invalid prediction",
        error="Pick 'real', 'synthetic', or 'unsure'.",
        promptTitle="Real or synthetic?",
        prompt="Use the dropdown: real / synthetic / unsure.",
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv_pred.add(f"B2:B{last_row}")
    ws.add_data_validation(dv_pred)

    dv_conf = DataValidation(
        type="list",
        formula1='"low,medium,high"',
        allow_blank=True,
        showDropDown=False,
        errorTitle="Invalid confidence",
        error="Pick 'low', 'medium', or 'high'.",
        promptTitle="How confident?",
        prompt="low / medium / high (optional).",
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv_conf.add(f"C2:C{last_row}")
    ws.add_data_validation(dv_conf)

    # ── Sheet 2: Instructions (auxiliary) ────────────────────────────────────
    ws_info = wb.create_sheet("Instructions")
    info_lines = [
        ("Reader study — real vs synthetic brain MRI", True, 14),
        ("", False, 11),
        (f"Total cases:  {n_real + n_synth} "
         f"({n_real} real + {n_synth} synthetic, randomly shuffled)", False, 11),
        ("", False, 11),
        ("Each case is in volumes/<case_id>/ and contains:", False, 11),
        ("    bravo.nii.gz — T1-weighted post-contrast MRI, normalized [0,1]", False, 11),
        ("    seg.nii.gz   — binary brain-metastasis segmentation", False, 11),
        ("", False, 11),
        ("How to fill the Reader Form:", True, 12),
        ("    1. Open each case in your viewer of choice.", False, 11),
        ("    2. Click the case_id row, then the prediction cell — use the dropdown.", False, 11),
        ("    3. (Optional) Set confidence: low / medium / high.", False, 11),
        ("    4. (Optional) Add comments explaining your reasoning.", False, 11),
        ("    5. Save the file as XLSX when finished.", False, 11),
        ("", False, 11),
        ("Allowed values:", True, 12),
        ("    prediction  →  real | synthetic | unsure", False, 11),
        ("    confidence  →  low | medium | high", False, 11),
        ("", False, 11),
        ("The answer key is in a separate file (answer_key.csv) and should not be", False, 11),
        ("consulted until after scoring.", False, 11),
    ]
    for i, (txt, bold, sz) in enumerate(info_lines, start=1):
        c = ws_info.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=sz)
        c.alignment = Alignment(vertical="center")
    ws_info.column_dimensions["A"].width = 95

    # Force the Reader Form to be the first/active sheet on open
    wb.active = wb.sheetnames.index("Reader Form")

    wb.save(str(path))
    logger.info("Wrote %s (Excel form with dropdowns + frozen header)", path)
    return True


def _process_case(
    src_dir: Path,
    dst_dir: Path,
    brain_mask: bool,
    threshold: float,
    dilate_pixels: int,
) -> None:
    """Load bravo + seg from src_dir, normalize+mask bravo, copy seg, save to dst_dir."""
    bravo_img = nib.load(str(src_dir / "bravo.nii.gz"))
    bravo = bravo_img.get_fdata().astype(np.float32)
    bravo = _normalize_bravo(bravo)
    if brain_mask:
        bravo = _apply_brain_mask(bravo, threshold=threshold, dilate_pixels=dilate_pixels)
    _save_nifti_like(bravo, bravo_img, dst_dir / "bravo.nii.gz")

    # Seg: copy unchanged to preserve label semantics.
    shutil.copy2(str(src_dir / "seg.nii.gz"), str(dst_dir / "seg.nii.gz"))


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--real-root", required=True, type=Path,
                        help="Real dataset root (e.g. .../brainmetshare-3).")
    parser.add_argument("--real-splits", nargs="+",
                        default=["train", "val", "test1", "test_new"],
                        help="Splits to union when building the real pool. "
                             "Default = all 4 splits (deduped → 156 unique patients).")
    parser.add_argument("--synth-dir", required=True, type=Path,
                        help="Synthetic root with <id>/{bravo,seg}.nii.gz cases.")
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--n-real", type=int, default=50)
    parser.add_argument("--n-synth", type=int, default=50)
    parser.add_argument("--seed", type=int, default=42)

    parser.add_argument("--brain-mask", action=argparse.BooleanOptionalAction, default=True,
                        help="Apply brain-mask zeroing to BOTH real and synth bravo "
                             "for fair comparison (default: yes).")
    parser.add_argument("--brain-mask-threshold", type=float, default=0.05)
    parser.add_argument("--brain-mask-dilate", type=int, default=2)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
    )
    args = parse_args(argv)
    rng = np.random.default_rng(args.seed)

    real_subjects = _find_real_subjects_across_splits(args.real_root, args.real_splits)
    synth_subjects = _find_subjects(args.synth_dir)
    logger.info("Real pool: %d unique subjects across splits %s under %s",
                len(real_subjects), args.real_splits, args.real_root)
    logger.info("Synth pool: %d subjects in %s", len(synth_subjects), args.synth_dir)

    if args.n_real > len(real_subjects):
        raise ValueError(
            f"Asked for {args.n_real} real but only {len(real_subjects)} available"
        )
    if args.n_synth > len(synth_subjects):
        raise ValueError(
            f"Asked for {args.n_synth} synth but only {len(synth_subjects)} available"
        )

    # Sample (without replacement) from each pool deterministically via seed.
    real_picks = [real_subjects[i] for i in rng.choice(
        len(real_subjects), size=args.n_real, replace=False)]
    synth_picks = [synth_subjects[i] for i in rng.choice(
        len(synth_subjects), size=args.n_synth, replace=False)]

    # Build the combined pool of (source_path, label) and shuffle to assign IDs.
    pool = [(p, "real") for p in real_picks] + [(p, "synthetic") for p in synth_picks]
    order = rng.permutation(len(pool))

    out_dir = args.output_dir
    volumes_dir = out_dir / "volumes"
    volumes_dir.mkdir(parents=True, exist_ok=True)
    id_width = len(str(len(pool)))  # zero-pad case IDs

    answer_rows: list[tuple[str, str, str]] = []
    case_ids: list[str] = []

    for new_idx, src_idx in enumerate(order, start=1):
        src_path, label = pool[src_idx]
        case_id = str(new_idx).zfill(id_width)
        dst_dir = volumes_dir / case_id

        logger.info("[%s] %s ← %s", case_id, label, src_path)
        _process_case(
            src_dir=src_path,
            dst_dir=dst_dir,
            brain_mask=args.brain_mask,
            threshold=args.brain_mask_threshold,
            dilate_pixels=args.brain_mask_dilate,
        )

        answer_rows.append((case_id, label, str(src_path)))
        case_ids.append(case_id)

    # ── Write answer key (KEEP HIDDEN) ──────────────────────────────────────
    answer_key_path = out_dir / "answer_key.csv"
    with open(answer_key_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["case_id", "true_label", "source_path"])
        w.writerows(answer_rows)
    logger.info("Wrote answer key: %s (DO NOT SHARE)", answer_key_path)

    # ── XLSX form with dropdowns (the reader's form) ────────────────────────
    xlsx_path = out_dir / "radiologist_form.xlsx"
    xlsx_ok = _write_xlsx_form(
        xlsx_path,
        case_ids=case_ids,
        n_real=args.n_real,
        n_synth=args.n_synth,
    )
    if not xlsx_ok:
        logger.error(
            "openpyxl is required to write the radiologist form. "
            "Install with: pip install openpyxl"
        )
        return 3

    # ── README explaining the form ──────────────────────────────────────────
    readme = out_dir / "README.txt"
    readme.write_text(
        "Reader study — real vs synthetic brain-MRI evaluation\n"
        "=====================================================\n"
        f"  {len(pool)} cases ({args.n_real} real + {args.n_synth} synthetic), randomly shuffled.\n"
        f"  Each case has bravo.nii.gz (T1-weighted post-contrast, normalized to [0,1])\n"
        "  and seg.nii.gz (binary brain-metastasis segmentation).\n\n"
        "Files in this directory:\n"
        f"  volumes/{1:0{id_width}d}..{len(pool):0{id_width}d}/  — case folders\n"
        "  radiologist_form.xlsx — Excel form with dropdowns (fill this in)\n"
        "  answer_key.csv        — TRUTH (do not open before reader scoring)\n\n"
        "Form columns:\n"
        "  case_id     — pre-filled, do not change\n"
        "  prediction  — 'real' or 'synthetic'\n"
        "  confidence  — low / medium / high  (optional)\n"
        "  comments    — free-text reasoning (optional)\n\n"
        f"Brain-mask zeroing applied to bravo: {args.brain_mask}\n"
    )
    logger.info("Wrote %s", readme)

    n_real_assigned = sum(1 for _, lbl, _ in answer_rows if lbl == "real")
    n_synth_assigned = sum(1 for _, lbl, _ in answer_rows if lbl == "synthetic")
    print(f"\nDone. {len(pool)} cases written to {volumes_dir}")
    print(f"  real: {n_real_assigned}   synthetic: {n_synth_assigned}")
    print(f"  answer key:        {answer_key_path}  (keep hidden)")
    print(f"  radiologist form:  {xlsx_path}  (share)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
